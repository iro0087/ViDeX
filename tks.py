import cv2, os, time, shutil

from termcolor import colored

from console_progressbar import ProgressBar

from openpyxl import load_workbook

l_frame = []

w = load_workbook("a.xlsx")

sheet = w.active

numb = str(sheet.cell(row=1, column=1).value)

sheet.cell(row=1, column=1).value = sheet.cell(row=1, column=1).value + 1

sheet.cell(row=1, column=2).value = sheet.cell(row=1, column=2).value + 1

cam = cv2.VideoCapture("tste.mp4")

data = "data" + numb

currentframe = 0

number_of_frames = int(cam.get(cv2.CAP_PROP_FRAME_COUNT))

fps = cam.get(cv2.CAP_PROP_FPS)

duration = number_of_frames / fps

fps2 = int(fps) + 1

nov = fps2

tot = fps2 * int(duration + 1)

fps = str(fps)

fps2 = str(fps2)

print("")

print("1s = " + fps + "frames = " + fps2 + "pages |  Total pages number = " + str(tot))

print("")

print("The video is at " + fps2 + "fps")  

print("")

lowr = str(input("Do you want to lower the number of pages wich will lower the fps video? (y/n)"))

print("")

passd_frame = 1

if lowr == "y":

    nov = int(input("At how many fps?" + "(modulo " + fps2 + ")"))

print("")

wantit = str(input("Proceed? (y/n)"))

print("")

if wantit == "y":

    if not os.path.exists(data):

        os.makedirs(data)

        print("creating folder "  + colored(data + "...", "red"))

        print("")

    while (True):

        ret, frame = cam.read()
            
        pb = ProgressBar(total=number_of_frames / (int(fps2)/nov),prefix='Here', suffix='Now', decimals=3, length=50, fill='#', zfill='-')

        if currentframe + 1 <= number_of_frames:

            pb.print_progress_bar(currentframe + 1)

        if ret:

            if int(fps2) / nov - passd_frame == 0:

                name = "./" + data + "/frame" + str(currentframe) + ".jpg"

                l_frame.append("frame" + str(currentframe) + ".jpg")

                cv2.imwrite(name, frame)

                currentframe += 1

                passd_frame = 0

            passd_frame += 1

        else:

            break

    cam.release()

    cv2.destroyAllWindows()

    w.save("a.xlsx")

    passd_frame = 0

    print("")

    print("")

    print("creating", colored("edit.tex...", "red"))

    print("")

    command = "touch " + data + "/edit.tex"

    os.system(command)

    command = "cat debut.tex >> " + data + "/edit.tex"

    filename_ = data + "/edit.tex" 

    os.system(command)

    while passd_frame + 1 < number_of_frames / (int(fps2) / nov):

        pb = ProgressBar(total=(number_of_frames - 1) / (int(fps2)/nov),prefix='Here', suffix='Now', decimals=3, length=50, fill='#', zfill='-')
        
        pb.print_progress_bar(passd_frame + 1)

        command = "bat mid.tex >> " + data + "/edit.tex"

        os.system(command)

        with open(filename_, "r") as file:
            
            filedata = file.read()

        filedata = filedata.replace("img.jpg", l_frame[passd_frame])

        with open(filename_, "w") as file:

            file.write(filedata)

        passd_frame += 1

    command = "cat mid2.tex >> " + data + "/edit.tex"

    os.system(command)

    with open(filename_, "r") as file:
            
            filedata = file.read()

    filedata = filedata.replace("img.jpg", l_frame[-1])

    with open(filename_, "w") as file:

            file.write(filedata)

    command = "cat fin.tex >> " + data + "/edit.tex"

    os.system(command)

    print("")

    print("done")

    print("")






